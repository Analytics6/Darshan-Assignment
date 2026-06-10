import json
import re
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
DATA_FILE = ROOT.parent / 'spacez_reviews_dataset (1).xlsx'
OUT_DIR = ROOT / 'data'
OUT_DIR.mkdir(exist_ok=True)

THEMES = [
    ('pool', r'pool'),
    ('cleanliness', r'clean|dirty|unclean|sheets|housekeeping|dish|damp'),
    ('check-in', r'check-in|check in|arrival|lat[e] arrival|late'),
    ('wifi', r'wifi|wi-fi'),
    ('heating', r'heat|heater|hot water|heating|geyser|AC|cold'),
    ('listing', r'photo|listing|misled|oversold|blocked|airport pickup|road|gate policy|occupancy policy|occupancy'),
    ('environment', r'mosquito|humidity|smell|damp|waterfront|backwaters|sunset|views'),
    ('policy', r'occupancy|gate policy|guests'),
    ('service', r'cook|houseboat|airport pickup|tips|chai|service|host|caretaker|manager|responsive|helpful|polite|organised|organized'),
]

OWNERSHIP = {
    'caretaker': {'check-in', 'service'},
    'operations': {'pool', 'cleanliness', 'wifi', 'heating', 'environment'},
    'business': {'listing', 'policy'},
}


def normalize_rating(raw, scale):
    try:
        raw = float(raw)
    except (TypeError, ValueError):
        return None
    if scale == 5:
        return raw
    if scale == 10:
        return raw / 2
    return raw


def assign_tags(text):
    tags = []
    for name, pattern in THEMES:
        if re.search(pattern, text, re.I):
            tags.append(name)
    if not tags:
        tags.append('general')
    return sorted(set(tags))


def ownership_for_tags(tags):
    owners = set()
    for owner, owner_tags in OWNERSHIP.items():
        if owner_tags.intersection(tags):
            owners.add(owner)
    if not owners:
        owners = {'operations'}
    return sorted(owners)


def sentiment_for_rating(norm):
    if norm is None:
        return 'unknown'
    if norm >= 4:
        return 'positive'
    if norm >= 3.5:
        return 'mixed'
    return 'negative'


def load_reviews():
    wb = load_workbook(DATA_FILE, data_only=True)
    sheet = wb['Reviews']
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    output = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        text = record.get('review_text') or ''
        norm_rating = normalize_rating(record.get('rating_raw'), record.get('rating_scale'))
        tags = assign_tags(text)
        owners = ownership_for_tags(tags)
        sentiment = sentiment_for_rating(norm_rating)
        record['rating_normalized'] = norm_rating
        record['tags'] = tags
        record['owners'] = owners
        record['sentiment'] = sentiment
        output.append(record)
    return output


def summarize(reviews):
    by_property = defaultdict(lambda: {'reviews': 0, 'avg_rating': 0, 'sentiment': defaultdict(int), 'tag_counts': defaultdict(int)})
    by_caretaker = defaultdict(lambda: {'reviews': 0, 'avg_rating': 0, 'sentiment': defaultdict(int), 'tag_counts': defaultdict(int)})

    for review in reviews:
        prop = review['property_name']
        caret = review['caretaker_name']
        norm = review['rating_normalized'] or 0

        for bucket, summary in (('property', by_property), ('caretaker', by_caretaker)):
            key = review['property_name'] if bucket == 'property' else review['caretaker_name']
            summary[key]['reviews'] += 1
            summary[key]['avg_rating'] += norm
            summary[key]['sentiment'][review['sentiment']] += 1
            for tag in review['tags']:
                summary[key]['tag_counts'][tag] += 1

    for summary in (by_property, by_caretaker):
        for key, data in summary.items():
            data['avg_rating'] = round(data['avg_rating'] / data['reviews'], 2)
            data['tag_counts'] = dict(sorted(data['tag_counts'].items(), key=lambda item: -item[1]))
            data['sentiment'] = dict(data['sentiment'])
    return {
        'properties': {k: v for k, v in by_property.items()},
        'caretakers': {k: v for k, v in by_caretaker.items()},
    }


def main():
    reviews = load_reviews()
    summary = summarize(reviews)
    output = {
        'reviews': reviews,
        'summary': summary,
    }
    with open(OUT_DIR / 'reviews.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, default=str)
    print(f'Wrote {OUT_DIR / "reviews.json"}')


if __name__ == '__main__':
    main()
